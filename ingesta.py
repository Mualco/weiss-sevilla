"""Convierte uno o varios Excel de estadisticas Weiss Sevilla en una SQLite limpia.

Uso:
    python ingesta.py archivo1.xlsx [archivo2.xlsx ...] [--db weiss.db]

Cada .xlsx corresponde a una temporada (p.ej. WeissSevilla-26.xlsx -> temporada 2026,
WeissSevilla-27.xlsx -> temporada 2027). La temporada de cada fila se deriva del anio
de su columna de fecha, no del nombre de archivo, asi que basta con pasar todos los
.xlsx juntos (el actual + los de anios anteriores) y la base se reconstruye entera
cada vez. No hay modo incremental: se recalcula todo desde los Excel fuente.

Solo se importan las hojas FUENTE (RAW RESULTS, Raw Data, Asistencia, Player Refs,
Set Refs, Deck Info). Las hojas derivadas (Ranking Temporada, Ranking mensual,
Torneos Recientes, TorneosMayo, Set_Stats, Deck Stats, Hoja 3, Resultados_Individuales)
se ignoran: la app las recalcula a partir de las tablas fuente.
"""

from __future__ import annotations

import argparse
import glob
import sqlite3
import sys
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import pandas as pd

REQUIRED_SHEETS = {
    "RAW RESULTS": ["ID", "Fecha", "Ronda", "Jugador", "Deck 1", "Arquetipo 1", "Rival", "Deck 2", "Arquetipo 2", "Resultado"],
    "Raw Data": ["FECHA", "Jugador", "Set", "Arquetipo", "Wins", "Losses", "ID TORNEO"],
    "Asistencia": ["ID", "Fecha", "Jugadores"],
    "Player Refs": ["Nombre", "Imagen"],
    "Set Refs": ["Codigo", "Imagen"],
    "Deck Info": ["Title", "Deck Archetype", "Image 1", "Image 2", "Image 3"],
}

VALID_RESULTS = {"W", "L"}


def clean_str_column(series: pd.Series) -> pd.Series:
    """Convierte a texto y recorta espacios, preservando los valores nulos como
    None (evita que pandas >= 3 convierta celdas vacias en el texto literal 'nan')."""
    return series.apply(lambda x: str(x).strip() if pd.notna(x) else None)


class IngestaError(Exception):
    """Error critico: el Excel no tiene el formato esperado."""


@dataclass
class Report:
    """Acumula avisos no criticos para mostrarlos juntos al final."""

    warnings: list[str] = field(default_factory=list)

    def warn(self, msg: str) -> None:
        self.warnings.append(msg)

    def flush(self) -> None:
        if not self.warnings:
            return
        print(f"\n{len(self.warnings)} aviso(s):")
        for w in self.warnings:
            print(f"  - {w}")


def find_header_row(ws, required_headers: list[str], max_scan: int = 10) -> tuple[int, dict[str, int]]:
    """Busca, en las primeras `max_scan` filas, la fila que contiene todas las
    cabeceras requeridas y devuelve (numero_de_fila, {cabecera: indice_columna})."""
    wanted = {h.strip().lower() for h in required_headers}
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        colmap: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            if isinstance(cell, str) and cell.strip().lower() in wanted:
                key = cell.strip()
                if key not in colmap:  # se queda con la primera aparicion (mas a la izquierda);
                    colmap[key] = col_idx  # cabeceras como "Jugador" se repiten en bloques derivados a la derecha
        found = {c.lower() for c in colmap}
        if wanted.issubset(found):
            # normaliza claves a como aparecen en required_headers, no como se escribieron
            norm = {}
            for h in required_headers:
                for k, v in colmap.items():
                    if k.lower() == h.strip().lower():
                        norm[h] = v
                        break
            return row_idx, norm
    raise IngestaError(
        f"No se encontro una fila con las cabeceras {required_headers} en la hoja "
        f"'{ws.title}'. El Excel puede haber cambiado de formato."
    )


def sheet_to_rows(ws, header_row: int, colmap: dict[str, int]) -> list[dict]:
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        record = {name: (row[idx] if idx < len(row) else None) for name, idx in colmap.items()}
        if all(v in (None, "") for v in record.values()):
            continue
        rows.append(record)
    return rows


def load_workbook_checked(source) -> openpyxl.Workbook:
    """`source` puede ser un Path o un objeto tipo-archivo (p.ej. UploadedFile)."""
    name = getattr(source, "name", str(source))
    try:
        wb = openpyxl.load_workbook(source, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise IngestaError(f"No se pudo abrir '{name}': {exc}") from exc
    missing = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
    if missing:
        raise IngestaError(
            f"'{name}' no tiene las hojas fuente esperadas: {missing}. "
            f"Hojas encontradas: {wb.sheetnames}"
        )
    return wb


def load_matches(wb, source: str, report: Report) -> pd.DataFrame:
    ws = wb["RAW RESULTS"]
    header_row, colmap = find_header_row(ws, REQUIRED_SHEETS["RAW RESULTS"])
    rows = sheet_to_rows(ws, header_row, colmap)
    df = pd.DataFrame(rows).rename(
        columns={
            "ID": "id_torneo",
            "Fecha": "fecha",
            "Ronda": "ronda",
            "Jugador": "jugador",
            "Deck 1": "set1",
            "Arquetipo 1": "arquetipo1",
            "Rival": "rival",
            "Deck 2": "set2",
            "Arquetipo 2": "arquetipo2",
            "Resultado": "resultado",
        }
    )
    total = len(df)
    df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
    bad_fecha = df["fecha"].isna().sum()
    if bad_fecha:
        report.warn(f"[{source}] matches: {bad_fecha} fila(s) con fecha no interpretable, descartadas.")
        df = df.dropna(subset=["fecha"])

    # descarta filas de ronda incompleta / bye: sin jugador o sin resultado
    keep = df["jugador"].notna() & df["jugador"].astype(str).str.strip().ne("") & df["resultado"].notna()
    dropped = total - int(keep.sum())
    if dropped:
        report.warn(f"[{source}] matches: {dropped} fila(s) sin jugador o sin resultado descartadas (rondas incompletas/byes).")
    df = df[keep].copy()

    bad_result = ~df["resultado"].astype(str).str.strip().str.upper().isin(VALID_RESULTS)
    if bad_result.any():
        report.warn(f"[{source}] matches: {int(bad_result.sum())} fila(s) con resultado distinto de W/L descartadas.")
        df = df[~bad_result]
    df["resultado"] = df["resultado"].astype(str).str.strip().str.upper()

    df["id_torneo"] = pd.to_numeric(df["id_torneo"], errors="coerce").astype("Int64")
    df["ronda"] = pd.to_numeric(df["ronda"], errors="coerce").astype("Int64")
    df["temporada"] = df["fecha"].dt.year
    for col in ("jugador", "rival", "set1", "arquetipo1", "set2", "arquetipo2"):
        df[col] = clean_str_column(df[col])

    dupes = df.duplicated(subset=["temporada", "id_torneo", "jugador", "ronda"]).sum()
    if dupes:
        report.warn(f"[{source}] matches: {dupes} fila(s) duplicadas (misma temporada/torneo/jugador/ronda).")

    return df[["temporada", "id_torneo", "fecha", "ronda", "jugador", "set1", "arquetipo1", "rival", "set2", "arquetipo2", "resultado"]]


def load_entries(wb, source: str, report: Report) -> pd.DataFrame:
    ws = wb["Raw Data"]
    header_row, colmap = find_header_row(ws, REQUIRED_SHEETS["Raw Data"])
    rows = sheet_to_rows(ws, header_row, colmap)
    df = pd.DataFrame(rows).rename(
        columns={
            "FECHA": "fecha",
            "Jugador": "jugador",
            "Set": "set",
            "Arquetipo": "arquetipo",
            "Wins": "wins",
            "Losses": "losses",
            "ID TORNEO": "id_torneo",
        }
    )
    # las filas de "Raw Data" que solo llevan bloques de resumen a la derecha
    # (sin Jugador) no son entries reales
    total = len(df)
    df = df[df["jugador"].notna() & df["jugador"].astype(str).str.strip().ne("")].copy()
    dropped = total - len(df)
    if dropped:
        report.warn(f"[{source}] entries: {dropped} fila(s) sin jugador descartadas.")

    df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
    bad_fecha = df["fecha"].isna().sum()
    if bad_fecha:
        report.warn(f"[{source}] entries: {bad_fecha} fila(s) con fecha no interpretable, descartadas.")
        df = df.dropna(subset=["fecha"])

    df["wins"] = pd.to_numeric(df["wins"], errors="coerce")
    df["losses"] = pd.to_numeric(df["losses"], errors="coerce")
    bad_num = df["wins"].isna() | df["losses"].isna() | (df["wins"] < 0) | (df["losses"] < 0)
    if bad_num.any():
        report.warn(f"[{source}] entries: {int(bad_num.sum())} fila(s) con wins/losses invalidos descartadas.")
        df = df[~bad_num]
    df["wins"] = df["wins"].astype(int)
    df["losses"] = df["losses"].astype(int)

    df["id_torneo"] = pd.to_numeric(df["id_torneo"], errors="coerce").astype("Int64")
    df["temporada"] = df["fecha"].dt.year
    for col in ("jugador", "set", "arquetipo"):
        df[col] = clean_str_column(df[col])

    dupes = df.duplicated(subset=["temporada", "id_torneo", "jugador", "set", "arquetipo"]).sum()
    if dupes:
        report.warn(f"[{source}] entries: {dupes} fila(s) duplicadas (misma temporada/torneo/jugador/set/arquetipo).")

    return df[["temporada", "id_torneo", "fecha", "jugador", "set", "arquetipo", "wins", "losses"]]


def load_attendance(wb, source: str, report: Report) -> pd.DataFrame:
    ws = wb["Asistencia"]
    header_row, colmap = find_header_row(ws, REQUIRED_SHEETS["Asistencia"])
    rows = sheet_to_rows(ws, header_row, colmap)
    df = pd.DataFrame(rows).rename(columns={"ID": "id_torneo", "Fecha": "fecha", "Jugadores": "jugadores"})
    df = df[df["id_torneo"].notna() & df["fecha"].notna()].copy()
    df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
    df = df.dropna(subset=["fecha"])
    df["jugadores"] = pd.to_numeric(df["jugadores"], errors="coerce").fillna(0).astype(int)
    df["id_torneo"] = pd.to_numeric(df["id_torneo"], errors="coerce").astype("Int64")
    df["temporada"] = df["fecha"].dt.year
    dupes = df.duplicated(subset=["temporada", "id_torneo"]).sum()
    if dupes:
        report.warn(f"[{source}] attendance: {dupes} fila(s) duplicadas (misma temporada/torneo).")
    return df[["temporada", "id_torneo", "fecha", "jugadores"]]


def load_player_refs(wb, source: str, report: Report) -> pd.DataFrame:
    ws = wb["Player Refs"]
    header_row, colmap = find_header_row(ws, REQUIRED_SHEETS["Player Refs"])
    rows = sheet_to_rows(ws, header_row, colmap)
    df = pd.DataFrame(rows).rename(columns={"Nombre": "jugador", "Imagen": "imagen"})
    df = df[df["jugador"].notna() & df["jugador"].astype(str).str.strip().ne("")].copy()
    df["jugador"] = clean_str_column(df["jugador"])
    return df[["jugador", "imagen"]]


def load_set_refs(wb, source: str, report: Report) -> pd.DataFrame:
    ws = wb["Set Refs"]
    header_row, colmap = find_header_row(ws, REQUIRED_SHEETS["Set Refs"])
    rows = sheet_to_rows(ws, header_row, colmap)
    df = pd.DataFrame(rows).rename(columns={"Codigo": "codigo", "Imagen": "imagen"})
    df = df[df["codigo"].notna() & df["codigo"].astype(str).str.strip().ne("")].copy()
    df["codigo"] = clean_str_column(df["codigo"])
    return df[["codigo", "imagen"]]


def load_deck_info(wb, source: str, report: Report) -> pd.DataFrame:
    ws = wb["Deck Info"]
    header_row, colmap = find_header_row(ws, REQUIRED_SHEETS["Deck Info"])
    rows = sheet_to_rows(ws, header_row, colmap)
    df = pd.DataFrame(rows).rename(
        columns={
            "Title": "set",
            "Deck Archetype": "arquetipo",
            "Image 1": "img1",
            "Image 2": "img2",
            "Image 3": "img3",
        }
    )
    df = df[df["set"].notna() & df["set"].astype(str).str.strip().ne("")].copy()
    df["set"] = clean_str_column(df["set"])
    df["arquetipo"] = clean_str_column(df["arquetipo"])
    return df[["set", "arquetipo", "img1", "img2", "img3"]]


def cross_check_references(matches: pd.DataFrame, entries: pd.DataFrame, player_refs: pd.DataFrame, set_refs: pd.DataFrame, report: Report) -> None:
    known_players = set(player_refs["jugador"])
    known_sets = set(set_refs["codigo"])

    def non_null(*series: pd.Series) -> set:
        values: set = set()
        for s in series:
            values |= set(s.dropna())
        return values

    unknown_players = non_null(matches["jugador"], entries["jugador"]) - known_players
    if unknown_players:
        report.warn(f"Jugadores sin foto en Player Refs (se mostraran sin imagen): {sorted(unknown_players)}")

    unknown_sets = non_null(matches["set1"], matches["set2"], entries["set"]) - known_sets
    if unknown_sets:
        report.warn(f"Codigos de set sin referencia en Set Refs (se mostraran sin imagen): {sorted(unknown_sets)}")


def build_database(files: list, db_path: Path) -> Report:
    """Reconstruye db_path desde `files`. Cada elemento puede ser un Path o
    cualquier objeto tipo-archivo con atributo `.name` (p.ej. un UploadedFile
    de Streamlit) para poder importar desde un formulario web."""
    report = Report()
    matches_parts, entries_parts, attendance_parts = [], [], []
    player_refs_parts, set_refs_parts, deck_info_parts = [], [], []

    for source in files:
        name = getattr(source, "name", str(source))
        print(f"Procesando {name}...")
        wb = load_workbook_checked(source)
        matches_parts.append(load_matches(wb, name, report))
        entries_parts.append(load_entries(wb, name, report))
        attendance_parts.append(load_attendance(wb, name, report))
        player_refs_parts.append(load_player_refs(wb, name, report))
        set_refs_parts.append(load_set_refs(wb, name, report))
        deck_info_parts.append(load_deck_info(wb, name, report))

    matches = pd.concat(matches_parts, ignore_index=True)
    entries = pd.concat(entries_parts, ignore_index=True)
    attendance = pd.concat(attendance_parts, ignore_index=True)

    # tablas de referencia: se combinan varias temporadas y se deduplican
    # quedandonos con la ultima aparicion (Excel mas reciente = mas fiable)
    player_refs = pd.concat(player_refs_parts, ignore_index=True).drop_duplicates(subset="jugador", keep="last")
    set_refs = pd.concat(set_refs_parts, ignore_index=True).drop_duplicates(subset="codigo", keep="last")
    deck_info = pd.concat(deck_info_parts, ignore_index=True).drop_duplicates(subset=["set", "arquetipo"], keep="last")

    cross_check_references(matches, entries, player_refs, set_refs, report)

    seasons = sorted(matches["temporada"].dropna().unique().tolist())
    print(f"Temporadas detectadas: {seasons}")
    print(f"matches={len(matches)} entries={len(entries)} attendance={len(attendance)} "
          f"player_refs={len(player_refs)} set_refs={len(set_refs)} deck_info={len(deck_info)}")

    report.flush()

    with sqlite3.connect(db_path) as conn:
        matches.to_sql("matches", conn, if_exists="replace", index=False)
        entries.to_sql("entries", conn, if_exists="replace", index=False)
        attendance.to_sql("attendance", conn, if_exists="replace", index=False)
        player_refs.to_sql("player_refs", conn, if_exists="replace", index=False)
        set_refs.to_sql("set_refs", conn, if_exists="replace", index=False)
        deck_info.to_sql("deck_info", conn, if_exists="replace", index=False)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_matches_temporada ON matches(temporada, id_torneo)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_entries_temporada ON entries(temporada, id_torneo)")

    print(f"OK: {db_path} generada.")
    return report


def expand_inputs(patterns: list[str]) -> list[Path]:
    paths: list[Path] = []
    for pattern in patterns:
        matches = sorted(glob.glob(pattern))
        if not matches:
            if Path(pattern).exists():
                matches = [pattern]
            else:
                raise IngestaError(f"No se encontro ningun archivo para '{pattern}'.")
        paths.extend(Path(m) for m in matches)
    # dedup preservando orden
    seen = set()
    unique = []
    for p in paths:
        if p not in seen:
            seen.add(p)
            unique.append(p)
    return unique


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("excel_files", nargs="+", help="Uno o varios .xlsx (uno por temporada). Admite patrones glob.")
    parser.add_argument("--db", default="weiss.db", help="Ruta de salida de la SQLite (por defecto weiss.db).")
    args = parser.parse_args(argv)

    try:
        files = expand_inputs(args.excel_files)
        build_database(files, Path(args.db))
    except IngestaError as exc:
        print(f"\nERROR: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
